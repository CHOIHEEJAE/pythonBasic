from openpyxl import Workbook


workbook = Workbook()
worksheet = workbook.active
worksheet.title ="sample_title"


worksheet["A1"] = 1
worksheet["A2"] = 2
worksheet["A3"] = 3
worksheet["A4"] = 4
worksheet["A5"] = 5
worksheet["A6"] = 6

print("A1 셀의 정보 출력 : ",worksheet["A1"])
print("A1 셀의 값 출력 : ",worksheet["A1"].value)
print("값이 없는 셀의 값을 출력하면 ? ", worksheet["A256"].value)

# 다른방법으로 셀에 접근하기
# column = A, B, C, ...  < 숫자로 접근
# row = 1, 2, 3, ...
cell_value = worksheet.cell(row = 7, column = 1, value = 7).value
print("다른방법으로 셀에 접근하기 : " ,cell_value)


# 반복문을 이용해 랜덤숫자 채워보기
from random import *

index = 1
for x in range(1,11) :
    for y in range(1, 11) :
        # worksheet.cell(column = x, row = y, value = randint(1, 100))
        worksheet.cell(column = x, row = y, value = index)
        index += 1



workbook.save("sample.xlsx")
