from openpyxl import load_workbook # 파일 불러오기

workbook = load_workbook("sample.xlsx") # 해당 파일 불러오기

worksheet = workbook.active
worksheet.title = "sample"

# 셀 값 가져오기
# for x in range(1, 11) :
#     for y in range(1, 11) :
#         print(worksheet.cell(column = x , row = y).value, end=" ")
#     print()

# cell 개수를 모를 때
for x in range(1, worksheet.max_column + 1) :
    for y in range(1, worksheet.max_row + 1) :
        # print(worksheet.cell(column = x , row = y).value, end=" ")
        print("column count : ",x)
        print("row count ", y)
    print()