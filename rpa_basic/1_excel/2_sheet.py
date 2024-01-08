from openpyxl import Workbook
workbook = Workbook()
worksheet = workbook.create_sheet() # 시트를 생성하면서 이름 설정하기, 빈값으로 생성하면 시트 기본이름으로 생성 (Sheet1, Sheet2 등)
worksheet.title = "newSheet1" # 이미 생성된 시트에 이름 바꾸기

worksheet.sheet_properties.tabColor = "ffffff"

worksheet1 = workbook.create_sheet("newSheet2")
worksheet2 = workbook.create_sheet("newSheet3", 2) # 생성된 시트를 해당 index 위치에 추가 (Sheet, newSheet1, newSheet3, newSheet2)

new_sheet = workbook["newSheet1"] # 시트 이름으로 해당 시트에 접근가능

print('모든 시트 title 확인 : ',workbook.sheetnames)

worksheet2["A1"] = "TEST_DATA" # newSheet3 A1 에 데이터 삽입
copied_sheet = workbook.copy_worksheet(worksheet2) # newSheet3 시트 복사
copied_sheet.title="copiedSheet" # 복사된 copied_sheet title 이름 설정

print('모든 시트 title 확인 : ',workbook.sheetnames)

workbook.save("sample.xlsx") # 파일명